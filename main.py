import datetime
import math
import random
import openpyxl
from sqlalchemy import create_engine
import models.models as model
from sqlalchemy.sql.expression import func
from sqlalchemy.orm import Session
from sqlalchemy import select
from faker import Faker
import os
from dotenv import load_dotenv

load_dotenv()

DB_USER = os.getenv('DB_USER')
DB_PASSWORD = os.getenv('DB_PASSWORD')
DB_HOST = os.getenv('DB_HOST')
DB_PORT = os.getenv('DB_PORT')
DB_NAME = os.getenv('DB_NAME')
DB_SCHEMA = os.getenv('DB_SCHEMA')

engine = create_engine(
    f'postgresql+psycopg2://{DB_USER}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}',
    connect_args={'options': '-csearch_path={}'.format(f'{DB_SCHEMA}')},
)
session = Session(engine)

# City fields
CITY_ID = 0
CITY_NAME = 1
CITY_LATITUDE = 2
CITY_LONGITUDE = 3

# Product fields
PRODUCT_ID = 0
PRODUCT_BRAND = 1
PRODUCT_MODEL = 2
PRODUCT_BODY_TYPE = 3
PRODUCT_YEAR = 4
PRODUCT_PRICE = 5

fake = Faker('id_ID')


def process_city_file():
    cities = []
    workbook = openpyxl.load_workbook('files/city.xlsx')
    worksheet = workbook['city']

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column,
                                   values_only=True):
        city = model.City(
            id=int(row[CITY_ID]),
            name=row[CITY_NAME],
            location='(' + str(row[CITY_LATITUDE]) + ',' + str(row[CITY_LONGITUDE]) + ')'
        )
        cities.append(city)

    session.add_all(cities)
    session.commit()


def process_product_file():
    products = []
    workbook = openpyxl.load_workbook('files/car_product.xlsx')
    worksheet = workbook['car_product']

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column,
                                   values_only=True):
        random_account = get_random_account()
        random_city = get_random_city()
        brand = insert_brand_if_not_exists(row[PRODUCT_BRAND])
        body_type = insert_type_if_not_exists(row[PRODUCT_BODY_TYPE])
        car_model = insert_model_if_not_exists(row[PRODUCT_MODEL], brand, body_type)

        products.append(model.Product(
            id=int(row[PRODUCT_ID]),
            year=int(row[PRODUCT_YEAR]),
            price=float(row[PRODUCT_PRICE]),
            model=car_model,
            account=random_account,
            city=random_city,
            created_at=create_new_datetime_within_dates(random_account.created_at)
        ))

    session.add_all(products)
    session.commit()


def insert_brand_if_not_exists(brand_name):
    q = session.query(model.Brand).filter(model.Brand.name == brand_name)
    result = session.query(q.exists())[0][0]
    brand = model.Brand()
    if not result:
        brand = model.Brand(
            name=brand_name
        )
        session.add(brand)
        session.commit()
    else:
        brand = session.execute(select(model.Brand).where(model.Brand.name == brand_name)).all()[0][0]

    return brand


def insert_type_if_not_exists(type_name):
    q = session.query(model.BodyType).filter(model.BodyType.name == type_name)
    result = session.query(q.exists())[0][0]
    body_type = model.BodyType()
    if not result:
        body_type = model.BodyType(
            name=type_name
        )
        session.add(body_type)
        session.commit()
    else:
        body_type = session.execute(select(model.BodyType).where(model.BodyType.name == type_name)).all()[0][0]

    return body_type


def insert_model_if_not_exists(model_name, brand, body_type):
    q = session.query(model.Model).filter(model.Model.name == model_name,
                                          model.Model.type_id == body_type.id,
                                          model.Model.brand_id == brand.id)
    result = session.query(q.exists())[0][0]
    car_model = model.Model()
    if not result:
        car_model = model.Model(
            name=model_name,
            type=body_type,
            brand=brand,
        )
        session.add(car_model)
        session.commit()
    else:
        car_model = session.execute(select(model.Model).where(model.Model.name == model_name,
                                                              model.Model.type_id == body_type.id,
                                                              model.Model.brand_id == brand.id)).all()[0][0]

    return car_model


def create_new_datetime_with_interval(start_datetime, interval):
    return fake.date_time_between(
        start_date=start_datetime,
        end_date=start_datetime + datetime.timedelta(minutes=interval),
    )


def create_new_datetime_within_dates(start_datetime, end_datetime=datetime.datetime.now()):
    return fake.date_time_between(
        start_date=start_datetime,
        end_date=end_datetime,
    )


def create_dummy_accounts(n):
    start_datetime = datetime.datetime(2022, 1, 1, 0, 0, 0, 0)

    last_account = session.query(model.Account).order_by(model.Account.id.desc()).first()

    if last_account is not None:
        start_datetime = last_account.created_at

    for i in range(n):
        random_city = get_random_city()
        new_datetime = create_new_datetime_with_interval(start_datetime, 7200)
        start_datetime = new_datetime
        session.add(model.Account(
            name=fake.name(),
            phone_number=fake.phone_number(),
            address=fake.address(),
            created_at=new_datetime,
            city=random_city,
        ))
        session.commit()


def get_random_account():
    account = session.query(model.Account).order_by(func.random()).first()
    return account


def get_random_city():
    city = session.query(model.City).order_by(func.random()).first()
    return city


def create_fake_location():
    return fake.local_latlng(country_code="ID", coords_only=True)


def create_dummy_bids(buyer_count_max, bid_precision_percentage, product=None):
    products = []
    if product == None:
        products = session.execute(select(model.Product).where(model.Product.deleted_at.is_(None))).all()
    else:
        products.append(product)
    for product in products:
        product = product[0]
        buyer_count = random.randint(0, buyer_count_max)
        for i in range(buyer_count):
            start_datetime = product.created_at
            price = math.ceil(product.price + (
                        product.price * random.uniform((0 - bid_precision_percentage), bid_precision_percentage)))
            last_bid = session.query(model.Bid).where(model.Bid.product_id == product.id).order_by(
                model.Bid.id.desc()).first()
            if last_bid is not None:
                start_datetime = last_bid.created_at
                price = math.ceil(last_bid.price +
                                  (last_bid.price * random.uniform((0 - bid_precision_percentage),
                                                                   bid_precision_percentage)))
            random_account = get_random_account()
            new_datetime = create_new_datetime_within_dates(start_datetime)
            session.add(model.Bid(
                account=random_account,
                product=product,
                price=price,
                status="sent",
                created_at=new_datetime,
            ))
            session.commit()


if __name__ == '__main__':
    process_city_file()
    create_dummy_accounts(100)
    process_product_file()
    create_dummy_bids(15, 0.05)

